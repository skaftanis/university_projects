import openpyxl
import numpy as np


def get_data_from_excel():

	#open xlsx file with openpyxl
	#original xls file doen't work with this lib, so I converted it to xlsx
	wb = openpyxl.load_workbook('BreastTissue.xlsx')
	sheet = wb.get_sheet_by_name('Data')

	classes=[]
	I0=[]
	PA500=[]
	HFS=[]
	DA=[]
	Area=[]
	ADA=[]
	MaxIP=[]
	DR=[]
	P=[]

	#take the data column by column
	for i in range(2,108):
		classes.append(sheet.cell(row=i, column=2).value)
		I0.append(sheet.cell(row=i, column=3).value)
		PA500.append(sheet.cell(row=i, column=4).value)
		HFS.append(sheet.cell(row=i, column=5).value)
		DA.append(sheet.cell(row=i, column=6).value)
		Area.append(sheet.cell(row=i, column=7).value)
		ADA.append(sheet.cell(row=i, column=8).value)
		MaxIP.append(sheet.cell(row=i, column=9).value)
		DR.append(sheet.cell(row=i, column=10).value)
		P.append(sheet.cell(row=i, column=11).value)

	return classes,I0,PA500,HFS,DA,Area,ADA,MaxIP,DR,P 


def get_normalized_data(): 
	#classes,I0,PA500,HFS,DA,Area,ADA,MaxIP,DR,P = get_data_from_excel()
	allData = get_data_from_excel()

	#print(np.mean(allData[5]))

	stats = []
	#create a list of dictionaries with avg,min and max for every data list (except for "classes")
	for i in range(1, len(allData) ):
		stats.append ({'avg': np.mean(allData[i]) , 'min': np.min(allData[i]) , 'max': np.max(allData[i]) })


	#calculate the normilized value of every data (except "classes")
	for i in range(1, len(allData) ):
		for j in range(len(allData[0])):
			allData[i][j] = ( allData[i][j] - stats[i-1]['avg'] ) / (  stats[i-1]['max'] - stats[i-1]['min']  )

	#convert class names with numbers 
	for j in range(len(allData[0])):
		if allData[0][j] == "car":
			allData[0][j] = 1
		elif allData[0][j] == "fad":
			allData[0][j] = 2
		elif allData[0][j] == "mas":
			allData[0][j] = 3
		elif allData[0][j] == "gla":
			allData[0][j] = 4 
		elif allData[0][j] == "con":
			allData[0][j] = 5
		elif allData[0][j] == "adi":
			allData[0][j] = 6

		

	return allData

#test the values (only if we call directly this script)
if __name__ == "__main__":
	#get_normalized_data()
	#classes,I0,PA500,HFS,DA,Area,ADA,MaxIP,DR,P = get_data_from_excel()
	classes,I0,PA500,HFS,DA,Area,ADA,MaxIP,DR,P = get_normalized_data()
	print(HFS)
	print("pa500")
	#print (PA500)